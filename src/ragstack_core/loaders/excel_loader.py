from collections.abc import Iterator
from pathlib import Path

import openpyxl

from ragstack_core.loaders.base_loader import BaseLoader
from ragstack_core.models.document_block import DocumentBlock
from ragstack_core.models.document_info import DocumentInfo


class ExcelLoader(BaseLoader):

    def __init__(self, rows_per_block: int = 100) -> None:
        if rows_per_block <= 0:
            raise ValueError("rows_per_block must be greater than 0")
        self._rows_per_block = rows_per_block

    def load_info(self, file_path: str | Path) -> DocumentInfo:
        return DocumentInfo.from_path(file_path)

    def load_blocks(
        self,
        file_path: str | Path,
        document_info: DocumentInfo,
    ) -> Iterator[DocumentBlock]:
        workbook = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        try:
            block_index = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows_iter = sheet.iter_rows(values_only=True)

                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    continue

                column_names = [str(cell) if cell is not None else "" for cell in header_row]
                block_rows: list[str] = []
                start_row = 1
                current_row = 0

                for row in rows_iter:
                    current_row += 1
                    block_rows.append(
                        " | ".join(
                            f"{col}: {'' if val is None else val}"
                            for col, val in zip(column_names, row)
                        )
                    )

                    if len(block_rows) >= self._rows_per_block:
                        yield self._create_block(
                            document_info=document_info,
                            block_index=block_index,
                            block_rows=block_rows,
                            column_names=column_names,
                            sheet_name=sheet_name,
                            start_row=start_row,
                            end_row=current_row,
                        )
                        block_index += 1
                        block_rows = []
                        start_row = current_row + 1

                if block_rows:
                    yield self._create_block(
                        document_info=document_info,
                        block_index=block_index,
                        block_rows=block_rows,
                        column_names=column_names,
                        sheet_name=sheet_name,
                        start_row=start_row,
                        end_row=current_row,
                    )
                    block_index += 1
        finally:
            workbook.close()

    def _create_block(
        self,
        document_info: DocumentInfo,
        block_index: int,
        block_rows: list[str],
        column_names: list[str],
        sheet_name: str,
        start_row: int,
        end_row: int,
    ) -> DocumentBlock:
        return DocumentBlock(
            document_id=document_info.document_id,
            block_index=block_index,
            text="\n".join(block_rows),
            metadata={
                "file_name": document_info.file_name,
                "file_type": document_info.file_type,
                "source_path": document_info.source_path,
                "block_type": "row_group",
                "rows_per_block": self._rows_per_block,
                "sheet_name": sheet_name,
                "start_row": start_row,
                "end_row": end_row,
                "column_names": column_names,
            },
        )
